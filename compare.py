
from process import image_match
from html_parse import get_test_case_name
import os
import re
import time
import sys
import openpyxl


def get_files(path):
    file_list=[]
    files= os.listdir(path)
    for f in files:
        if f.startswith("ReRun1") and f.endswith(".html"):
            file_list.append(f)

    return file_list

def get_first_run_files(path):
    first_run_file_list=[]
    files=os.listdir(path)
    for f in files:
        if f.startswith("INFO") and f.endswith(".html"):
            first_run_file_list.append(f)

    return first_run_file_list
    

def find_images():
    start_time=time.clock()
    
    path2=os.getcwd()
    
    suite_list=[]
    
    #path=raw_input("Enter the path in which results are present: ")
    path=sys.argv[1]
    os.chdir(path)
    wb1=openpyxl.Workbook()
    ws1=wb1.active
    ws1.cell(row=1,column=1).value="SuiteName"
    ws1.cell(row=1,column=2).value="TestCaseName"
    ws1.title="Sync issues"
    
    wb2=openpyxl.Workbook()
    ws2=wb2.active
    ws2.cell(row=1,column=1).value="SuiteName"
    ws2.cell(row=1,column=2).value="TestCaseName"
    ws2.title="Hard Errors"
    
    sheet_rows=2
    sheet_rows_h=2
    
    file_name=sys.argv[2]
    #file_name=raw_input("Enter the name of the regression analyis file (.xlsx): ")
    path_sync=os.path.join(path2,"templates")
    path_hard_error=os.path.join(path2,"hard_error")
    file_list=get_files(path)

    first_run_file_list=get_first_run_files(path)
    

    
    for f in file_list:
        #print "file being processed: ",f
        final_list=[]
        case_list=[]
        hard_error_list=[]
        final_hard_error_list=[]
        test_suite,data=get_test_case_name(f)
        #print "data from get_test_case_name",data

        for p in first_run_file_list:
            test_suite_first_run,data_first_run=get_test_case_name(p)

            if test_suite_first_run==test_suite:
                
                
            
        
        for i in data:
            for j in data[i]:
                match= image_match(j,path,path_sync)
                hard_error_match=image_match(j,path,path_hard_error)
                if match:
                    case_list.append(i)
                    
                if hard_error_match:
                    hard_error_list.append(i)
                
        for x in case_list:
            if x not in final_list:
                final_list.append(x)
                ws1.cell(row=sheet_rows,column=1).value=test_suite
                ws1.cell(row=sheet_rows,column=2).value=x
                sheet_rows+=1
                print "Sync Issue: "+test_suite+"\t"+x+"\n"
        for y in hard_error_list:
            if y not in final_hard_error_list:
                final_hard_error_list.append(y)
                ws2.cell(row=sheet_rows_h,column=1).value=test_suite
                ws2.cell(row=sheet_rows_h,column=2).value=y
                sheet_rows_h+=1
                print "Hard Error: "+test_suite+"\t"+y+"\n"

    



    wb=openpyxl.load_workbook(file_name)
    ws=wb.active
    i1=3  #start of the suite names in the analysis file
    i2=2  #start of the suite names in the generated file.

    for p in range(i1,ws.max_row+1):
        
        
            
        if ws.cell(row=p,column=1).value==None:
            break
        else:
            
            
            for q in range(i2,ws1.max_row+1):
                
                if (ws1.cell(row=q,column=1).value in ws.cell(row=p,column=1).value) and (ws.cell(row=p,column=2).value==ws1.cell(row=q,column=2).value):
                    ws.cell(row=p,column=5).value="Stuck in Processing"
                    ws.cell(row=p,column=6).value="Sync Issue"
            for r in range(i2,ws2.max_row+1):
                
                if (ws2.cell(row=r,column=1).value in ws.cell(row=p,column=1).value) and (ws.cell(row=p,column=2).value==ws2.cell(row=r,column=2).value):
                    ws.cell(row=p,column=5).value="Hard Error"
                    ws.cell(row=p,column=6).value="App Issue"
                   
        
        
    wb.save(file_name)
    



    
    
    print "Please check the result analysis file"
    end_time=time.clock()
    print "total time: ",end_time-start_time



if __name__=='__main__':
    find_images()
       

            
        
    



